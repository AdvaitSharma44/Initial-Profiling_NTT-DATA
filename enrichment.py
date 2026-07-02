from __future__ import annotations

import io
import json
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import httpx
import pandas as pd
from fastapi import UploadFile
from openpyxl.styles import PatternFill

from backend.models.config import AppConfig
from backend.models.schemas import CompanyData, RawCompanyProfile

logger = logging.getLogger("lead_enrichment_api.services")
config = AppConfig()

MOCK_INTERNAL_ARCHIVE: Dict[Tuple[str, str], Dict] = {
    ("future pharma pvt ltd", "mumbai"): {
        "industry": "Pharmaceuticals",
        "sub_sector": "Pharma",
        "revenue_inr_crores": 120.5,
        "employee_count": 240,
        "status": "Active",
        "parent_company": None,
        "data_freshness_year": 2025,
        "cagr": 18.3,
    },
    ("bright health solutions", "delhi"): {
        "industry": "Healthcare",
        "sub_sector": "Hospital Services",
        "revenue_inr_crores": 52.1,
        "employee_count": 95,
        "status": "Active",
        "parent_company": None,
        "data_freshness_year": 2024,
        "cagr": 12.0,
    },
}


def validate_upload_file_extension(filename: str) -> bool:
    return str(filename).strip().lower().endswith(tuple(config.allowed_extensions))


def is_allowed_content_type(content_type: str) -> bool:
    if not content_type:
        return True
    content_type = content_type.split(";")[0].strip().lower()
    return content_type in config.allowed_content_types


async def validate_upload_file(file: UploadFile) -> Tuple[str, bytes]:
    filename = file.filename or ""
    extension = Path(filename).suffix.lower()
    if not validate_upload_file_extension(filename):
        raise ValueError("Invalid file extension")

    content_type = (file.content_type or "").split(";")[0].strip().lower()
    if content_type and not is_allowed_content_type(content_type):
        raise ValueError("Invalid MIME type")

    max_bytes = config.max_upload_size_mb * 1024 * 1024
    payload = await file.read(max_bytes + 1)
    if len(payload) > max_bytes:
        raise ValueError("File exceeds maximum upload size")
    return extension, payload


def parse_dataframe(file_bytes: bytes, extension: str) -> pd.DataFrame:
    try:
        buffer = io.BytesIO(file_bytes)
        if extension == ".xlsx":
            df = pd.read_excel(buffer, engine="openpyxl")
        else:
            df = pd.read_csv(buffer)
        if len(df) > config.max_rows:
            raise ValueError("Row count exceeds maximum allowed")
        return df
    except Exception as exc:
        logger.error("Failed to parse uploaded file: %s", exc)
        raise


def check_existing_data(row: pd.Series) -> bool:
    existing_value = row.get("Industry") or row.get("industry")
    if existing_value is None or not pd.notna(existing_value):
        return False
    normalized = str(existing_value).strip()
    if not normalized:
        return False
    if normalized.lower() in {"n/a", "na", "not available", "unknown", "none"}:
        return False
    return True


def check_internal_archive(company_name: str, region: str) -> Optional[RawCompanyProfile]:
    key = (company_name.strip().lower(), region.strip().lower())
    archived = MOCK_INTERNAL_ARCHIVE.get(key)
    if archived:
        return RawCompanyProfile(**archived, region=region)
    return None


def normalize_number(value: str) -> Optional[float]:
    if not value:
        return None
    cleaned = value.replace(",", "").replace(" ", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def convert_to_crores(amount: float, unit: str) -> Optional[float]:
    unit_lower = unit.strip().lower()
    if unit_lower in {"crore", "crores", "cr"}:
        return round(amount, 2)
    if unit_lower in {"million", "mn", "m"}:
        return round(amount / 10.0, 2)
    if unit_lower in {"billion", "bn"}:
        return round(amount * 100.0, 2)
    return None


async def mock_external_api(company_name: str, region: str) -> Optional[RawCompanyProfile]:
    if not config.apollo_api_key:
        print(f"DEBUG: Apollo API key missing for {company_name}")
        logger.warning("Apollo API key is not configured; skipping Apollo enrichment for %s.", company_name)
        return RawCompanyProfile(region=region)

    endpoint = "https://api.apollo.io/v1/organizations/search"
    headers = {
        "X-Api-Key": config.apollo_api_key,
        "Accept": "application/json",
    }
    payload = {
        "q_organization_name": company_name,
    }
    print(f"DEBUG: Apollo request for {company_name} with headers {headers} and payload {payload}")

    data = None
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            response = await client.post(endpoint, json=payload, headers=headers)
            print(f"DEBUG: Status Code for {company_name}: {response.status_code}")
            print(f"DEBUG: Response text for {company_name}: {response.text[:2000]}")
            response.raise_for_status()
            data = response.json()
    except Exception as exc:
        print(f"DEBUG: CRITICAL FAILURE for {company_name} - {exc}")
        logger.warning("Apollo enrichment request failed for %s: %s", company_name, exc)
        return RawCompanyProfile(region=region)

    print(f"DEBUG: Response JSON for {company_name}: {data}")

    def extract_organizations(response_data: Any) -> List[Dict[str, Any]]:
        if isinstance(response_data, list):
            return [item for item in response_data if isinstance(item, dict)]
        if not isinstance(response_data, dict):
            return []
        if isinstance(response_data.get("organizations"), list):
            return [item for item in response_data["organizations"] if isinstance(item, dict)]
        if isinstance(response_data.get("results"), list):
            return [item for item in response_data["results"] if isinstance(item, dict)]
        if isinstance(response_data.get("data"), dict):
            return extract_organizations(response_data["data"])
        if isinstance(response_data.get("data"), list):
            return [item for item in response_data["data"] if isinstance(item, dict)]
        if "organization" in response_data and isinstance(response_data["organization"], dict):
            return [response_data["organization"]]
        if all(key in response_data for key in ("name", "city")):
            return [response_data]
        return []

    organizations = extract_organizations(data)
    if not organizations:
        print(f"DEBUG: No organizations parsed from Apollo response for {company_name}")
        if isinstance(data, dict):
            print(f"DEBUG: Apollo response keys: {list(data.keys())}")
        return RawCompanyProfile(region=region)

    organization = organizations[0]
    print(f"DEBUG: Parsed org for {company_name}: {organization}")

    def parse_numeric_value(value: Any) -> Optional[float]:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.replace(",", "").strip()
            if not cleaned:
                return None

            # Parse revenue strings with magnitude suffixes like 10M, 5B, 200K, 50 Cr
            match = re.search(r"([-+]?[0-9]*\.?[0-9]+)\s*([kmbt]|million|billion|crore|cr|k|m|b|t)?", cleaned, re.IGNORECASE)
            if match:
                number_str = match.group(1)
                suffix = (match.group(2) or "").lower()
                try:
                    number = float(number_str)
                except ValueError:
                    return None

                if suffix in {"k", "thousand"}:
                    return number * 1_000
                if suffix in {"m", "million"}:
                    return number * 1_000_000
                if suffix in {"b", "billion"}:
                    return number * 1_000_000_000
                if suffix in {"t", "trillion"}:
                    return number * 1_000_000_000_000
                if suffix in {"cr", "crore"}:
                    return number * 10_000_000

                return number

            return None
        return None

    def extract_nested_numeric(value: Any) -> Optional[float]:
        if isinstance(value, dict):
            for key in ("amount", "value", "revenue", "annual_revenue", "estimated_revenue", "low", "high", "min", "max", "display"):
                candidate = value.get(key)
                parsed = parse_numeric_value(candidate)
                if parsed is not None:
                    return parsed
            return None
        return parse_numeric_value(value)

    industries = organization.get("industries")
    if isinstance(industries, list) and industries:
        industry = str(industries[0]).strip()
    else:
        industry = "Data Unavailable"

    revenue_usd = None
    revenue_source = None
    for revenue_key in (
        "organization_revenue",
        "annual_revenue",
        "estimated_annual_revenue",
        "revenue",
        "revenue_usd",
        "annual_revenue_usd",
        "estimated_revenue",
        "revenue_range",
        "organization_revenue_printed",
    ):
        revenue_value = organization.get(revenue_key)
        revenue_usd = extract_nested_numeric(revenue_value)
        if revenue_usd is not None:
            revenue_source = revenue_key
            print(f"DEBUG: Found revenue value for {company_name} via '{revenue_key}': {revenue_value} -> {revenue_usd}")
            break

    emp_val = (
        organization.get("estimated_num_employees")
        or organization.get("employee_count")
        or organization.get("num_employees")
        or organization.get("employees")
    )
    emp_parsed = extract_nested_numeric(emp_val)
    employee_count = int(emp_parsed) if emp_parsed is not None else None

    discovered_location = organization.get("city")
    discovered_location = str(discovered_location).strip() if discovered_location else None
    final_region = discovered_location if discovered_location else region

    status = organization.get("status") or organization.get("company_status") or None
    status = str(status).strip() if status else "Data Unavailable"

    parent_company = organization.get("parent_company") or organization.get("parent") or None
    parent_company = str(parent_company).strip() if parent_company else "Data Unavailable"

    return RawCompanyProfile(
        industry=industry,
        revenue_usd=revenue_usd,
        revenue_inr_crores=None,
        employee_count=employee_count,
        status=status,
        parent_company=parent_company,
        data_freshness_year=None,
        history_revenue_crores={},
        region=final_region,
    )


def convert_usd_to_inr_crores(usd_amount: Optional[float]) -> Optional[float]:
    if usd_amount is None:
        return None
    amount = float(usd_amount)
    if amount <= 0:
        return None
    # Convert from USD to INR crores using an exchange rate of ~83 INR per USD.
    return round(amount * 83.0 / 10_000_000.0, 2)


def calculate_cagr(revenue_data: Dict[int, float]) -> Optional[float]:
    if not revenue_data or len(revenue_data) < 2:
        return None
    years = sorted(revenue_data.keys())
    start_value = revenue_data[years[0]]
    end_value = revenue_data[years[-1]]
    num_years = years[-1] - years[0]
    if start_value <= 0 or num_years <= 0:
        return None
    return round((pow(end_value / start_value, 1 / num_years) - 1) * 100, 2)


def map_industry_taxonomy(raw_industry: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not raw_industry:
        return None, None
    raw_lower = raw_industry.strip().lower()
    # If Apollo already provides a structured category (e.g. 'airlines/aviation'), prefer it
    if "/" in raw_lower or "&" in raw_lower:
        return raw_industry.title(), None

    # Prefer exact or strong keyword matches mapped to our verticals
    for vertical, keywords in {
        "Pharma": ["pharmaceutical", "drug", "medicine", "biotech", "pharma"],
        "Healthcare": ["hospital", "clinic", "health", "medical", "diagnostic", "healthcare"],
        "BFSI": ["bank", "finance", "insurance", "brokerage", "nbfc", "financial"],
        "IT/ITeS": ["software", " it ", "technology", "data", "cloud", "ai", "digital", "tech"],
        "Services": ["consulting", "audit", "legal", "marketing", "advertising", "service"],
        "Aviation": ["airline", "airlines", "aviation", "airliner"],
        "Manufacturing": ["manufacturing", "industrial", "factory"],
    }.items():
        if any(keyword in raw_lower for keyword in keywords):
            return vertical, None
    if "manufacturing" in raw_lower or "industrial" in raw_lower or "factory" in raw_lower:
        for subsector, keywords in {
            "Automobile": ["auto", "vehicle", "car", "truck", "motorcycle"],
            "Electrical Appliances": ["appliance", "electrical", "white goods"],
            "Machinery": ["machine", "equipment", "industrial", "machinery"],
            "Chemicals": ["chemical", "plastic", "petrochemical"],
        }.items():
            if any(keyword in raw_lower for keyword in keywords):
                return "Manufacturing", subsector
        return "Manufacturing", "General Manufacturing"
    if "health" in raw_lower or "clinic" in raw_lower or "hospital" in raw_lower:
        return "Healthcare", None
    if "bank" in raw_lower or "finance" in raw_lower or "insurance" in raw_lower:
        return "BFSI", None
    if "software" in raw_lower or "technology" in raw_lower or "data" in raw_lower:
        return "IT/ITeS", None
    if "consult" in raw_lower or "service" in raw_lower or "marketing" in raw_lower:
        return "Services", None
    return raw_industry.title(), None


def process_hierarchy_and_status(profile: RawCompanyProfile) -> RawCompanyProfile:
    normalized_status = profile.status.strip().title() if profile.status else "Active"
    profile.status = normalized_status
    if normalized_status in {"Acquired", "Subsidiary"} and not profile.parent_company:
        profile.parent_company = "Unknown Parent"
    return profile


def generate_confidence_tag(profile: RawCompanyProfile) -> str:
    status = (profile.status or "").title()
    if status in {"Bankrupt", "Written Off", "Defunct"}:
        return "[FLAG: Inactive Entity]"
    if status in {"Acquired", "Subsidiary"}:
        return "[Acquired - See Parent]"
    if (profile.revenue_inr_crores is None or profile.revenue_inr_crores < 10) and (profile.employee_count or 0) > 50:
        return "[POTENTIAL WHALE - Requires Manual Triage]"
    if profile.revenue_inr_crores is not None and profile.industry:
        return "[Ready for Sales]"
    return "[Requires Manual Review]"


def build_company_data(company_name: str, region: str, profile: RawCompanyProfile, original_row: pd.Series) -> CompanyData:
    if not profile.industry and check_existing_data(original_row):
        profile.industry = str(original_row.get("Industry") or original_row.get("industry")).strip()
    if profile.revenue_usd is not None and profile.revenue_inr_crores is None:
        profile.revenue_inr_crores = convert_usd_to_inr_crores(profile.revenue_usd)
    if profile.revenue_inr_crores is None and profile.history_revenue_crores:
        latest_year = max(profile.history_revenue_crores.keys())
        profile.revenue_inr_crores = profile.history_revenue_crores[latest_year]
        profile.data_freshness_year = profile.data_freshness_year or latest_year
    profile.cagr = profile.cagr or calculate_cagr(profile.history_revenue_crores)
    profile = process_hierarchy_and_status(profile)
    if profile.industry:
        profile.industry, profile.sub_sector = map_industry_taxonomy(profile.industry)
    return CompanyData(
        company_name=company_name,
        region=profile.region or region,
        industry=profile.industry,
        sub_sector=profile.sub_sector,
        revenue_inr_crores=profile.revenue_inr_crores,
        employee_count=profile.employee_count,
        status=profile.status,
        parent_company=profile.parent_company,
        cagr=profile.cagr,
        data_freshness_year=profile.data_freshness_year,
        confidence_tag=generate_confidence_tag(profile),
    )


def companies_to_excel_payload(companies: List[CompanyData]) -> bytes:
    rows = []
    for company in companies:
        row = {key: (value if value is not None else "N/A") for key, value in company.dict().items()}
        rows.append(row)
    dataframe = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Enriched Leads")
        worksheet = writer.sheets["Enriched Leads"]
        header_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
        for cell in worksheet[1]:
            cell.fill = header_fill
        tag_column_index = next((idx for idx, cell in enumerate(worksheet[1], start=1)
                                 if cell.value == "confidence_tag" or cell.value == "Confidence Tag" or cell.value == "Confidence & Action Tag"),
                                None)
        if tag_column_index:
            for row in worksheet.iter_rows(min_row=2, min_col=tag_column_index, max_col=tag_column_index):
                cell = row[0]
                value = str(cell.value or "").lower()
                if "ready for sales" in value:
                    cell.fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
                elif "acquired" in value:
                    cell.fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
                elif "inactive" in value or "bankrupt" in value:
                    cell.fill = PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid")
                elif "whale" in value:
                    cell.fill = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")
    output.seek(0)
    return output.getvalue()


